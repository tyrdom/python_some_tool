import os
import sys
import matplotlib.pyplot as plt
import xlrd
from openpyxl import load_workbook

res_dir = 'configXlsx'
in_file = sys.path[0] + os.sep + res_dir + os.sep + 'line_res.xlsm'
out_file = sys.path[0] + os.sep + res_dir + os.sep + 'line.xlsx'

r_sheet = 'bezier'
w_sheet = 'out'

data = xlrd.open_workbook(in_file)
table = data.sheet_by_name(r_sheet)
start_row = 3
start_col = 3
r_len = 25

wb = load_workbook(out_file)
_sth = wb.active
ws = wb[w_sheet]


def a_round(some_float):
    return int(round(some_float * 64)) / 64


def bezier(_p0, _p1, _p2, t):
    x0 = _p0[0]
    x1 = _p1[0]
    x2 = _p2[0]
    xt = (1 - t) * (1 - t) * x0 + 2 * t * (1 - t) * x1 + t * t * x2
    y0 = _p0[1]
    y1 = _p1[1]
    y2 = _p2[1]
    yt = (1 - t) * (1 - t) * y0 + 2 * t * (1 - t) * y1 + t * t * y2
    return [xt, yt]


def distance(pa, pb):
    return ((pb[0] - pa[0]) ** 2 + (pb[1] - pa[1]) ** 2) ** 0.5


def direction(pa, pb):
    return [(pb[0] - pa[0]) / distance(pa, pb), (pb[1] - pa[1]) / distance(pa, pb)]


def join_p(a, b):
    # print('b' + str(b))
    return a + b[1:]


def gen_a_piece_i(a_p0, a_p1, a_p2):
    def gen_mid(_t0, _t1):
        return bezier(a_p0, a_p1, a_p2, (_t0 + _t1) / 2)

    def gen_a_p(l_p, r_p, t_l, t_r):
        f = distance(l_p, r_p)
        if f < r_len:
            return [l_p, r_p + [f / speed]]
        else:
            mid_p = gen_mid(t_l, t_r)
            t_m = (t_l + t_r) / 2
            l_part = gen_a_p(l_p, mid_p, t_l, t_m)

            r_part = gen_a_p(mid_p, r_p, t_m, t_r)

            return join_p(l_part, r_part)

    return gen_a_p(a_p0, a_p2, 0, 1)


for x in range(start_row - 1, table.nrows):
    line_id = table.row_values(x, 0, 1)[0]
    speed = table.row_values(x, start_col - 2, start_col - 1)[0]

    a_row = table.row_values(x, start_col - 1, table.ncols)
    a_data = list(filter(lambda s: s != '', a_row))

    out = []
    if a_data:
        print('speed~~~~~~', speed)
        p_data = [a_data[i:i + 3] for i in range(0, len(a_data), 3)]

        ar_time = distance(p_data[0], p_data[1]) / speed
        out.extend(
            [[a_round(p_data[0][0]), a_round(p_data[0][1]), 0.0],
             [a_round(p_data[1][0]), a_round(p_data[1][1]), a_round(ar_time)]])

        last_p1 = [p_data[0][0], p_data[0][1]]
        for i in range(1, len(p_data) - 1):
            offset = p_data[i][2]
            p0 = [p_data[i][0], p_data[i][1]]
            now_dir = direction(last_p1, p0)
            p1 = [p0[0] + offset * now_dir[0], p0[1] + offset * now_dir[1]]
            last_p1 = p1

            p2 = [p_data[i + 1][0], p_data[i + 1][1]]
            piece_i = gen_a_piece_i(p0, p1, p2)[1:]

            for q in range(len(piece_i)):
                ar_time += piece_i[q][2]
                piece_i[q] = [a_round(piece_i[q][0]), a_round(piece_i[q][1]), a_round(ar_time)]
            out.extend(piece_i)

        out_put = str(out).replace(', ', '=')[1:-1]
        print('out:::::', line_id, out_put)

        outputx = []
        outputy = []
        for i in range(len(out)):
            outputx.append(out[i][0])
            outputy.append(out[i][1])
        fig, axs = plt.subplots()
        axs.plot(outputx, outputy)
        plt.show()

        ws.cell(x, 1, out_put)
        cd = a_round(out[-1][2])
        ws.cell(x, 2, cd)
wb.save(out_file)
